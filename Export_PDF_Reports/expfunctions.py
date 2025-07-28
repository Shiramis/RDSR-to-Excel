import pandas as pd
from sqlalchemy import create_engine
import json
import os
import openpyxl
from openpyxl.utils import get_column_letter
from Axiom_process import make_excel

class expfun:
    def export_to_sql(pdf_files):
        """Export DataFrames to PostgreSQL database."""
        if not pdf_files:
            print("No PDF files to process.")
            return

        engine = create_engine(os.getenv("DB_URL"))

        for index, pdf_file in enumerate(pdf_files, start=1):
            data = make_excel()
            df_st, df_fl, dft, indiv, dfin, person, evst, evfl = data.startpro(pdf_file.file_path, index)

            # Καθαρισμός και προετοιμασία
            if df_st is not None and not df_st.empty:
                df_st['patient_index'] = index
                df_st.to_sql("dsa_reports", engine, if_exists="append", index=False)

            if df_fl is not None and not df_fl.empty:
                df_fl['patient_index'] = index
                df_fl.to_sql("fluoro_reports", engine, if_exists="append", index=False)

            if dft is not None and not dft.empty:
                dft['patient_index'] = index
                dft.to_sql("dose_summary", engine, if_exists="append", index=False)

            if dfin is not None and not dfin.empty:
                dfin['patient_index'] = index
                dfin.to_sql("study_info", engine, if_exists="append", index=False)

            print(f"Exported Patient {index} to database.")


    def export_to_json_fhir(df: pd.DataFrame, dfper: pd.DataFrame, file_path: str):
        """FHIR-compatible Bundle JSON with Patient + Observations."""

        if df is None or df.empty or dfper is None or dfper.empty:
            return

        # Extract patient info from dfper (it is transposed)
        person = dfper[0].tolist()  # Αν columns=[“”] και index είναι το όνομα του πεδίου

        patient_resource = {
            "resourceType": "Patient",
            "id": person[1],   # Patient ID
            "name": [{"text": person[0]}],
            "gender": person[2].lower(),
            "extension": [
                {
                    "url": "http://hl7.org/fhir/StructureDefinition/patient-age",
                    "valueAge": {"value": int(person[3]), "unit": "years"}
                }
            ]
        }

        entry = [{"resource": patient_resource}]

        records = df.to_dict(orient="records")
        for i, record in enumerate(records):
            obs_resource = {
                "resourceType": "Observation",
                "id": f"obs-{i}",
                "status": "final",
                "subject": {"reference": f"Patient/{person[1]}"},
                "value": record
            }
            entry.append({"resource": obs_resource})

        bundle = {
            "resourceType": "Bundle",
            "type": "collection",
            "entry": entry
        }

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(bundle, f, ensure_ascii=False, indent=2)

    def export_to_excel(excel_output_file, pdf_files, patients, individual, d, index, nst, nfl):
        """Export DataFrame to Excel file."""
        if not pdf_files:
            print("No PDF files to process.")
            return
        with pd.ExcelWriter(excel_output_file, engine='openpyxl') as writer:
            for pdf_file in pdf_files:

                data = make_excel()

                df_st, df_fl, dft, indiv, dfin, person, evst, evfl = data.startpro(pdf_file.file_path, index)

                dft.replace(0, "empty", inplace=True)
                dfper = pd.DataFrame(person,
                                    index=['Patient Name', 'Patient ID', 'Gender', 'Age (years)', 'Study Type', 'Manufacturer',
                                            'Content Date', 'Content Time', 'Person Observer Name',
                                            'Number of irradiation events'], columns=[""])
                
                start_row_dfper = 0
                start_row_df = start_row_dfper + len(dfper) + 2
                if df_st is not None and not df_st.empty:
                    nst += 1
                    dfper.loc['Number of irradiation events'] = evst
                    df_st.replace(0, "empty", inplace=True)
                    sheet_name_df_st = f"DSA-Patient {index}"
                    df_styled = df_st.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
                    df1_st = df_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
                    df1_st.to_excel(writer, sheet_name=sheet_name_df_st, startrow=start_row_df)

                    person[0] = f"Patient {index}"
                    person[1] = f"Patient ID {index}"
                    person[-2] = f"Observer {index}"

                    dfper_styled = dfper.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
                    df2 = dfper_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
                    df2.to_excel(writer, sheet_name=sheet_name_df_st, startrow=start_row_dfper, header=False)

                if df_fl is not None and not df_fl.empty:
                    nfl += 1
                    dfper.loc['Number of irradiation events'] = evfl
                    df_fl.replace(0, "empty", inplace=True)
                    sheet_name_df_fl = f"Fluoro-Patient {index}"
                    df_styled = df_fl.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
                    df1_fl = df_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
                    df1_fl.to_excel(writer, sheet_name=sheet_name_df_fl, startrow=start_row_df)

                    person[0] = f"Patient {index}"
                    person[1] = f"Patient ID {index}"
                    person[-2] = f"Observer {index}"

                    dfper_styled = dfper.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
                    df2 = dfper_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
                    df2.to_excel(writer, sheet_name=sheet_name_df_fl, startrow=start_row_dfper, header=False)

                patients.append(dft)
                individual.append(dfin)

                if df_st is not None and not df_st.empty:
                    d["sheet_st {0}".format(str(nst))] = sheet_name_df_st
                    d["max_row_st {0}".format(str(nst))] = max(start_row_df + len(df_st), start_row_dfper + len(dfper))
                    d["max_col_st {0}".format(str(nst))] = len(df_st.columns) + 1
                if df_fl is not None and not df_fl.empty:
                    d["sheet_fl {0}".format(str(nfl))] = sheet_name_df_fl
                    d["max_row_fl {0}".format(str(nfl))] = max(start_row_df + len(df_fl), start_row_dfper + len(dfper))
                    d["max_col_fl {0}".format(str(nfl))] = len(df_fl.columns) + 1
                index +=1
                print(f"Processed Patient {index}... ")
            
            dfpat = pd.concat([df for df in patients if not df.empty], axis=0)
            dfstpat = dfpat.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            dfpat = dfstpat.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
            dfpat.to_excel(writer, sheet_name=f"Accumulated X-Ray Dose Data")
            dfind = pd.concat([df for df in individual if not df.empty ], axis=0)
            dfstind = dfind.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            dfindiv = dfstind.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
            dfindiv.to_excel(writer, sheet_name=f"Basic Study Information")
        #---------------------------------------------------------------------------------------------------
        wb = openpyxl.load_workbook(excel_output_file)
        # Style and format the Excel sheets
        # Set the row height for the first row in each sheet
        # For each DSA sheet
        for i in range(1, nst + 1):
            sheet = wb[d["sheet_st " + str(i)]]
            max_row = d["max_row_st " + str(i)]
            max_col = d["max_col_st " + str(i)]
            sheet.row_dimensions[start_row_df+1].height = 31

            for row in range(start_row_df+1, max_row + 2):
                for col in range(1, max_col + 1):
                    sheet.cell(row=start_row_df + 1, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
                    sheet.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')

                column_widths = [24.9, 22.5, 16.5, 13, 16.5, 19, 19, 14, 11, 11, 19, 9.5, 10.5, 12, 13.5, 14, 14, 15.1, 15.3,
                                15.9, 19.3]
                for i, width in enumerate(column_widths, start=1):
                    column_letter = get_column_letter(i)
                    sheet.column_dimensions[column_letter].width = width
        #--------------------------------------------------------------------------------------------------            
        # For each Fluoro sheet            
        for i in range(1, nfl + 1):
            sheet = wb[d["sheet_fl " + str(i)]]
            max_row = d["max_row_fl " + str(i)]
            max_col = d["max_col_fl " + str(i)]
            sheet.row_dimensions[start_row_df+1].height = 31

            for row in range(start_row_df+1, max_row + 2):
                for col in range(1, max_col + 1):
                    sheet.cell(row=start_row_df + 1, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
                    sheet.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')

                column_widths = [24.9, 22.5, 16.5, 13, 16.5, 19, 19, 14, 11, 11, 19, 9.5, 10.5, 12, 13.5, 14, 14, 15.1, 15.3,
                                15.9, 19.3]
                for i, width in enumerate(column_widths, start=1):
                    column_letter = get_column_letter(i)
                    sheet.column_dimensions[column_letter].width = width
        #--------------------------------------------------------------------------------------------------            
        # Set the row height for the summary sheets
        sheet1 = wb["Accumulated X-Ray Dose Data"]
        sheet2 = wb["Basic Study Information"]
        sheet1.row_dimensions[1].height = 31
        sheet2.row_dimensions[1].height = 31
        alignment_settings = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')
        for col in range(1, 12):
            cell1 = sheet1.cell(row=1, column=col)
            cell1.alignment = alignment_settings
            cell2 = sheet2.cell(row=1, column=col)
            cell2.alignment = alignment_settings
        column_widths = [12, 13, 16.1, 9, 18.9, 13, 11, 19, 15, 15, 14]
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet1.column_dimensions[column_letter].width = width
        column_widths = [11.1, 12.8, 7, 10, 25, 12.3, 12, 11.9, 14.5]
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet2.column_dimensions[column_letter].width = width
        # Move the summary sheets to the start of the workbook
        wb.move_sheet(sheet1,offset=-nst-nfl)
        wb.move_sheet(sheet2, offset=-nst-nfl-1)
        # Save the workbook
        wb.save(excel_output_file)
        print(f"Excel output saved to: {excel_output_file}")
        # Return the DataFrames for further processing if needed
        df_st = df_st.reset_index(drop=True) if df_st is not None else pd.DataFrame()
        df_fl = df_fl.reset_index(drop=True) if df_fl is not None else pd.DataFrame()  
#-------------------------------------------------------------------------------------------------