import openpyxl
from Cios_process import make_excel
import os
import pandas as pd
from openpyxl.utils import get_column_letter

class PDFFile:
    def __init__(self, filename, file_path):
        self.filename = filename
        self.file_path = file_path

f = input (r'Write the path of the folder with DATA: ').strip('"')
folder_path = f
pdf_files = []
for filename in os.listdir(folder_path):
    if filename.endswith(".pdf"):
        file_path = os.path.join(folder_path, filename)
        pdf_file = PDFFile(filename, file_path)
        pdf_files.append(pdf_file)
ot = input(r"Write the path of the excel file: ").strip('"')
output_directory = ot
patients = []
individual = []
index = 0
d = {}

with pd.ExcelWriter(output_directory, engine='openpyxl') as writer:
    for pdf_file in pdf_files:
        data = make_excel()
        df, dft, indiv, dfin, person, name_id1, name_id2 = data.startpro(f"{pdf_file.file_path}",index)
        df.replace(0, "empty", inplace=True)
        dft.replace(0, "empty", inplace=True)
        sheet_name_df = f"Patient {index}"
        person[0] = f"Patient {index}"
        person[1] = f"Patient ID {index}"
        person[-2] = f"Observer {index}"
        print (person[0])
        dfper = pd.DataFrame(person,
                                  index=['Patient Name', 'Patient ID','Gender','Age (years)',
                                         'Study Type', 'Manufacturer', 'Content Date', 'Content Time',
                                         'Person Observer Name','Number of irradiation events'], columns=[""])
        start_row_dfper = 0
        start_row_df = start_row_dfper + len(dfper) + 2
        # Apply style to dataframes
        dfper_styled = dfper.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
        df_styled = df.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
        df1 = df_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
        df2 = dfper_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})

        df2.to_excel(writer, sheet_name=sheet_name_df, startrow=start_row_dfper, header=False)
        df1.to_excel(writer, sheet_name=sheet_name_df, startrow=start_row_df)

        patients.append(dft)
        individual.append(dfin)
        d["sheet {0}".format(str(index))] = sheet_name_df
        d["max_row {0}".format(str(index))] = max(start_row_df + len(df), start_row_dfper + len(dfper))
        d["max_col {0}".format(str(index))] = len(df.columns) + 1

        index +=1
    dfpat = pd.concat([df for df in patients if not df.empty], axis=0)
    dfstpat = dfpat.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
    dfpat = dfstpat.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
    dfpat.to_excel(writer, sheet_name=f"Accumulated X-Ray Dose Data")
    dfind = pd.concat([df for df in individual if not df.empty], axis=0)
    dfstind = dfind.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
    dfindiv = dfstind.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
    dfindiv.to_excel(writer, sheet_name=f"Basic Study Indormation")

wb = openpyxl.load_workbook(ot)

for i in range(0, index):
    sheet = wb[d["sheet " + str(i)]]
    max_row = d["max_row " + str(i)]
    max_col = d["max_col " + str(i)]
    sheet.row_dimensions[start_row_df+1].height = 31

    for row in range(start_row_df+1, max_row + 2):
        for col in range(1, max_col + 1):
            sheet.cell(row=start_row_df + 1, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
            sheet.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')
        column_widths = [24.9, 22.5, 16.5, 13, 16.5, 19, 19, 14, 11, 11, 19, 9.5, 10.5, 12, 13.5, 14, 14, 15.1, 15.3,
                         15.9, 19.3]
        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width

sheet1 = wb["Accumulated X-Ray Dose Data"]
sheet2 = wb["Basic Study Indormation"]
sheet1.row_dimensions[1].height = 31
sheet2.row_dimensions[1].height = 31
alignment_settings = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')
for col in range(1, 12):
    cell1 = sheet1.cell(row=1, column=col)
    cell1.alignment = alignment_settings
    cell2 = sheet2.cell(row=1, column=col)
    cell2.alignment = alignment_settings
column_widths = [12, 13, 16.1, 9, 18.9, 13, 11, 19, 15, 15, 14]
for i, width in enumerate(column_widths, start=1):
    column_letter = get_column_letter(i)
    sheet1.column_dimensions[column_letter].width = width
column_widths = [11.1, 12.8, 7, 10, 25, 12.3, 12, 11.9, 14.5]
for i, width in enumerate(column_widths, start=1):
    column_letter = get_column_letter(i)
    sheet2.column_dimensions[column_letter].width = width


wb.move_sheet(sheet1,offset=-index)
wb.move_sheet(sheet2, offset=-index)

wb.save(ot)

def check_and_rename_sheet(writer, sheet_name):
    while sheet_name in writer.sheets:
        base_sheet_name, counter = sheet_name, 1
        while True:
            new_sheet_name = f"{base_sheet_name} ({counter})"
            if new_sheet_name not in writer.sheets:
                return new_sheet_name
            counter += 1
