import os
import re
import glob
import time
import tkinter as tk
from tkinter import filedialog

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =========================
# Helpers
# =========================
def sanitize_path(p: str) -> str:
    if not p:
        return ""
    return os.path.normpath(p)


def clean_cell(x):
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x)).strip()


def safe_float(s):
    if s is None:
        return None
    s = clean_cell(s).replace(",", ".")
    m = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", s)
    return float(m.group(0)) if m else None


def extract_patient_name(page_text: str):
    m = re.search(r"Patient Name:\s*(.+)", page_text or "", flags=re.IGNORECASE)
    return clean_cell(m.group(1)) if m else None


def _parse_table_rows(rows):
    out = {}
    for r in rows[1:]:
        rr = [x for x in r if x != ""]
        if len(rr) >= 2:
            organ = rr[0]
            dose = safe_float(rr[1])
            if dose is None:
                dose = clean_cell(rr[1]) or None
            if organ:
                out[organ] = dose
    return out


def _parse_scan_rows(rows):
    out = {}
    for r in rows[1:]:
        rr = [x for x in r if x != ""]
        if len(rr) >= 2:
            key = rr[0]
            display_key, kind = _normalize_scan_param_name(key)
            if kind == "float":
                value = safe_float(rr[1])
            else:
                value = clean_cell(rr[1]) or None
            if display_key:
                out[display_key] = value
    return out


def _normalize_scan_param_name(name):
    raw = clean_cell(name)
    lowered = raw.lower()

    mapping = [
        ("ray direction", ("Ray Direction", "text")),
        ("rotational angle", ("Rotational Angle", "float")),
        ("tilt angle", ("Tilt Angle", "float")),
        ("longitudinal field of view", ("Longitudinal FOV (cm)", "float")),
        ("longitudinal fov", ("Longitudinal FOV (cm)", "float")),
        ("cross-body field of view", ("Cross-body FOV (cm)", "float")),
        ("cross body field of view", ("Cross-body FOV (cm)", "float")),
        ("cross-body fov", ("Cross-body FOV (cm)", "float")),
        ("cross body fov", ("Cross-body FOV (cm)", "float")),
        ("tube voltage", ("Tube Voltage (kVp)", "float")),
        ("filter cu", ("Filter Cu (mm)", "float")),
        ("source-to-imager", ("SID (cm)", "float")),
        ("source to imager", ("SID (cm)", "float")),
        ("source-to-skin", ("SSD (cm)", "float")),
        ("source to skin", ("SSD (cm)", "float")),
    ]

    for token, result in mapping:
        if token in lowered:
            return result

    return raw, "text"

def _parse_position_output_value(rows):
    header_row = None
    header_idx = None
    for i, r in enumerate(rows):
        lowered = [clean_cell(c).lower() for c in r]
        if any("position" in c for c in lowered) and any("output" in c for c in lowered) and any("value" in c for c in lowered):
            header_row = lowered
            header_idx = i
            break
    if header_row is None:
        return {}

    def find_idx(token):
        for i, c in enumerate(header_row):
            if token in c:
                return i
        return None

    pos_idx = find_idx("position")
    out_idx = find_idx("output")
    val_idx = find_idx("value")
    if pos_idx is None and out_idx is None and val_idx is None:
        return {}

    for r in rows[header_idx + 1:]:
        def get(idx):
            if idx is None or idx >= len(r):
                return ""
            return clean_cell(r[idx])

        position = get(pos_idx)
        output = get(out_idx)
        value = get(val_idx)
        if any([position, output, value]):
            result = {}
            if pos_idx is not None:
                result["Position"] = position or None
            if out_idx is not None:
                result["Output"] = output or None
            if val_idx is not None:
                result["Value"] = value or None
            return result

    return {}


def extract_dose_tables_from_pdf(pdf_path):
    """
    Returns:
      patient_label (str): Patient Name if found, else filename stem
      organ_doses (dict): { organ_name: dose_float }
      remainder_doses (dict): { organ_name: dose_float }
      scan_params (dict): { param_name: param_value }
      scan_extra (dict): { Position/Output/Value }
    """
    organ_doses = {}
    remainder_doses = {}
    scan_params = {}
    scan_extra = {}
    patient_label = os.path.splitext(os.path.basename(pdf_path))[0]

    with pdfplumber.open(pdf_path) as pdf:
        page1 = pdf.pages[0]
        text = page1.extract_text() or ""

        pname = extract_patient_name(text)
        if pname:
            patient_label = pname
        else:
            patient_label = "Patient name not found"

        tables = page1.extract_tables() or []
        for tbl in tables:
            rows = [[clean_cell(c) for c in row] for row in (tbl or []) if row]
            rows = [r for r in rows if any(r)]
            if len(rows) < 2:
                continue

            header = " ".join(rows[0]).lower()
            if ("organ" in header and "dose" in header):
                if not organ_doses:
                    organ_doses = _parse_table_rows(rows)
            elif "remainder" in header and "organ" in header:
                if not remainder_doses:
                    remainder_doses = _parse_table_rows(rows)
            elif "scan parameters" in header:
                if not scan_params:
                    scan_params = _parse_scan_rows(rows)

        if len(pdf.pages) > 1:
            page2 = pdf.pages[1]
            tables2 = page2.extract_tables() or []
            for tbl in tables2:
                rows = [[clean_cell(c) for c in row] for row in (tbl or []) if row]
                rows = [r for r in rows if any(r)]
                if len(rows) < 2:
                    continue

                if not scan_extra:
                    scan_extra = _parse_position_output_value(rows)
                if scan_extra:
                    break

    return patient_label, organ_doses, remainder_doses, scan_params, scan_extra


# =========================
# Main (your UI flow)
# =========================

# open a window
root = tk.Tk()
root.withdraw()

# Select PDF folder
folder_path = filedialog.askdirectory(title="Select Folder with PDF Dose Reports")
folder_path = sanitize_path(folder_path)

# Select output folder for the Excel file
initial_output_dir = os.path.dirname(folder_path)
output_folder = filedialog.askdirectory(
    title="Select Folder to Save the Excel File",
    initialdir=initial_output_dir
)
output_folder = sanitize_path(output_folder)

# Automatically create an Excel file path
timestamp = time.strftime("%Y-%m-%d_%H%M%S")
output_filename = f"Dose_Report_{timestamp}.xlsx"
output_directory = os.path.join(output_folder, output_filename)

# Collect PDFs
pdf_files = sorted(glob.glob(os.path.join(folder_path, "*.pdf")))
if not pdf_files:
    raise RuntimeError(f"No PDF files found in folder: {folder_path}")
start_time = time.time()

# ===== Create ONE workbook (from scratch) =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "OrganDoses"

# ===== Setup constants =====
# Row 2 = patient headers starting from col C
# Col B = organ names starting from row 3
HEADER_ROW = 2
ORGAN_COL = 2           # B
FIRST_PATIENT_COL = 3   # C
FIRST_ORGAN_ROW = 3

ws.cell(HEADER_ROW, ORGAN_COL).value = "Organ Dose"

SCAN_PARAM_ORDER = [
    "Ray Direction",
    "Rotational Angle",
    "Tilt Angle",
    "Longitudinal FOV (cm)",
    "Cross-body FOV (cm)",
    "Tube Voltage (kVp)",
    "Filter Cu (mm)",
    "SID (cm)",
    "SSD (cm)",
]

organ_order = []
remainder_order = []
scan_order = list(SCAN_PARAM_ORDER)
scan_extra_order = []
patient_data = []
next_patient_col = FIRST_PATIENT_COL

# ===== Fill workbook =====
for pdf_path in pdf_files:
    try:
        patient_label, doses, remainder_doses, scan_params, scan_extra = extract_dose_tables_from_pdf(pdf_path)
        patient_data.append((patient_label, doses, remainder_doses, scan_params, scan_extra))

        for organ in doses.keys():
            if organ not in organ_order:
                organ_order.append(organ)

        for organ in remainder_doses.keys():
            if organ not in remainder_order:
                remainder_order.append(organ)

        for param in scan_params.keys():
            if param not in scan_order:
                scan_order.append(param)

        for param in ["Position", "Output", "Value"]:
            if param in scan_extra and param not in scan_extra_order:
                scan_extra_order.append(param)
    except Exception as e:
        print(f"[ERROR] {os.path.basename(pdf_path)}: {e}")

remainder_label_row = FIRST_ORGAN_ROW + len(organ_order) + 1
remainder_header_row = remainder_label_row + 1
last_remainder_row = remainder_header_row + len(remainder_order) - 1

scan_label_row = last_remainder_row + 2
scan_start_row = scan_label_row + 1
scan_extra_start_row = scan_start_row + len(scan_order)

ws.cell(remainder_label_row, ORGAN_COL).value = "Remainder Organs"
ws.cell(scan_label_row, ORGAN_COL).value = "Scan Parameters"

for idx, organ in enumerate(organ_order):
    ws.cell(FIRST_ORGAN_ROW + idx, ORGAN_COL).value = organ

for idx, organ in enumerate(remainder_order):
    ws.cell(remainder_header_row + idx, ORGAN_COL).value = organ

for idx, param in enumerate(scan_order):
    ws.cell(scan_start_row + idx, ORGAN_COL).value = param

for idx, param in enumerate(scan_extra_order):
    ws.cell(scan_extra_start_row + idx, ORGAN_COL).value = param

for i, (_, doses, remainder_doses, scan_params, scan_extra) in enumerate(patient_data):
    col = next_patient_col + i
    patient_number = col - FIRST_PATIENT_COL + 1
    ws.cell(HEADER_ROW, col).value = f"Patient {patient_number}"
    ws.cell(remainder_header_row - 1, col).value = f"Patient {patient_number}"
    ws.cell(scan_label_row, col).value = f"Patient {patient_number}"

    for organ, dose in doses.items():
        r = FIRST_ORGAN_ROW + organ_order.index(organ)
        ws.cell(r, col).value = dose

    for organ, dose in remainder_doses.items():
        r = remainder_header_row + remainder_order.index(organ)
        ws.cell(r, col).value = dose

    for param, value in scan_params.items():
        r = scan_start_row + scan_order.index(param)
        ws.cell(r, col).value = value

    for param, value in scan_extra.items():
        r = scan_extra_start_row + scan_extra_order.index(param)
        ws.cell(r, col).value = value

    print(f"Patient {patient_number} exported")

# ===== Styling =====
last_col = FIRST_PATIENT_COL + len(patient_data) - 1
last_scan_row = scan_start_row + len(scan_order) - 1
if scan_extra_order:
    last_scan_row = scan_extra_start_row + len(scan_extra_order) - 1
last_row = max(last_remainder_row, scan_label_row, last_scan_row)

header_fill = PatternFill("solid", fgColor="D9D9D9")
section_fill = PatternFill("solid", fgColor="D9D9D9")
highlight_fill = PatternFill("solid", fgColor="DDEBF7")
bold_font = Font(bold=True)
center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="000000")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

# Header rows
for col in range(ORGAN_COL, last_col + 1):
    cell = ws.cell(HEADER_ROW, col)
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = left if col == ORGAN_COL else center

for col in range(ORGAN_COL, last_col + 1):
    cell = ws.cell(remainder_header_row - 1, col)
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = left if col == ORGAN_COL else center

for col in range(ORGAN_COL, last_col + 1):
    cell = ws.cell(scan_label_row, col)
    cell.fill = header_fill
    cell.font = bold_font
    cell.alignment = left if col == ORGAN_COL else center


# Scan extra rows (Position/Output/Value)
for idx, _ in enumerate(scan_extra_order):
    row = scan_extra_start_row + idx
    for col in range(ORGAN_COL, last_col + 1):
        cell = ws.cell(row, col)
        cell.fill = section_fill
        if col == ORGAN_COL:
            cell.font = bold_font

# Highlight summary rows in main section
summary_labels = {"peak skin dose", "effective dose"}
for organ in organ_order:
    lower_name = organ.lower()
    if any(label in lower_name for label in summary_labels):
        row = FIRST_ORGAN_ROW + organ_order.index(organ)
        for col in range(ORGAN_COL, last_col + 1):
            cell = ws.cell(row, col)
            cell.fill = highlight_fill
            if col == ORGAN_COL:
                cell.font = bold_font

# Borders
for r in range(HEADER_ROW, last_row + 1):
    for c in range(ORGAN_COL, last_col + 1):
        ws.cell(r, c).border = border

# Column widths
ws.column_dimensions["B"].width = 34.5
for col in range(FIRST_PATIENT_COL, last_col + 1):
    col_letter = openpyxl.utils.get_column_letter(col)
    ws.column_dimensions[col_letter].width = 20

# Save ONE file
wb.save(output_directory)
print(f"Saved: {output_directory}")
elapsed = time.time() - start_time
print(f"Processed PDFs: {len(pdf_files)}")
print(f"Elapsed time: {elapsed:.2f} seconds")
