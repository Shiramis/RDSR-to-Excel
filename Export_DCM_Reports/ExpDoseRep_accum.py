import tkinter as tk
from tkinter import filedialog
import pydicom
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime as dt
import time

class DicomFile:
    def __init__(self, filename, file_path):
        self.filename = filename
        self.file_path = file_path

def sanitize_path(path):
    return path.strip('"')

def extract_and_format_age(age_value):
    """Extract the numeric part of the age and format it."""
    if isinstance(age_value, str):
        numeric_part = ''.join(filter(str.isdigit, age_value))
        if numeric_part:
            return int(numeric_part)
        else:
            raise ValueError(f"No numeric part found in value '{age_value}'")

def extract_data(dicom_data):
    #==Totals===
    DAPtotal = []  # Dose Area Product Total
    RPt = []  # Dose (RP) Total
    dstrp = []  # Distance Source to Reference Point
    fDAPt = []  # Fluoro Dose Area Product Total
    fRPt = []  # Fluoro Dose (RP) Total
    tftime = []  # Total Fluoro Time
    aDAPt = []  # Acquisition Dose Area Product Total
    aRPt = []  # Acquisition Dose (RP) Total
    rpd = []  # Reference Point Definition
    tatime = []  # Total Acquisition Time

    def search_sequence(sequence):
        global count  # Ensure the 'count' variable is accessible globally
        count = False  # Initialize 'count' at the start of the sequence search

        for item in sequence:
            # Check if the item has the Concept Name Code Sequence (0040,A043)
            if (0x0040, 0xA043) in item:
                concept_name_code_sequence = item[(0x0040, 0xA043)].value
                code_meaning = concept_name_code_sequence[0].get((0x0008, 0x0104), None)
                if code_meaning:
                    code_meaning_value = code_meaning.value
                    if code_meaning_value == 'Dose Area Product Total':
                        extract_numeric_value(item, (0x0040, 0xA300), DAPtotal)
                    elif code_meaning_value == 'Dose (RP) Total':
                        extract_numeric_value(item, (0x0040, 0xA300), RPt)
                    elif code_meaning_value == 'Distance Source to Reference Point':
                        extract_numeric_value(item, (0x0040, 0xA300), dstrp)
                    elif code_meaning_value == 'Fluoro Dose Area Product Total':
                        extract_numeric_value(item, (0x0040, 0xA300), fDAPt)
                    elif code_meaning_value == 'Fluoro Dose (RP) Total':
                        extract_numeric_value(item, (0x0040, 0xA300), fRPt)
                    elif code_meaning_value == 'Total Fluoro Time':
                        extract_numeric_value(item, (0x0040, 0xA300), tftime)
                    elif code_meaning_value == 'Acquisition Dose Area Product Total':
                        extract_numeric_value(item, (0x0040, 0xA300), aDAPt)
                    elif code_meaning_value == 'Acquisition Dose (RP) Total':
                        extract_numeric_value(item, (0x0040, 0xA300), aRPt, allow_empty=True)
                    elif code_meaning_value == 'Reference Point Definition':
                        extract_concept_code(item, (0x0040, 0xA168), rpd)
                    elif code_meaning_value == 'Total Acquisition Time':
                        extract_numeric_value(item, (0x0040, 0xA300), tatime)
            # Recursively search within nested sequences
            if (0x0040, 0xA730) in item:
                search_sequence(item[(0x0040, 0xA730)].value)

    # Function to extract numeric value from Measured Value Sequence
    def extract_numeric_value(item, tag, target_list, allow_empty=False):
        if tag in item:
            measured_value_sequence = item[tag].value
            if measured_value_sequence and (0x0040, 0xA30A) in measured_value_sequence[0]:
                numeric_value = measured_value_sequence[0][(0x0040, 0xA30A)].value
                target_list.append(float(numeric_value))
            elif allow_empty:
                target_list.append('empty')
            else:
                target_list.append("N/A")

    # Function to extract code value from Concept Code Sequence
    def extract_concept_code(item, tag, target_list):
        if tag in item:
            concept_code_sequence = item[tag].value
            if concept_code_sequence and (0x0008, 0x0104) in concept_code_sequence[0]:
                code_value = concept_code_sequence[0][(0x0008, 0x0104)].value
                target_list.append(code_value)
            else:
                target_list.append('N/A')
    # Start searching from the main sequence
    if (0x0040, 0xA730) in dicom_data:
        search_sequence(dicom_data[(0x0040, 0xA730)].value)

    return DAPtotal, RPt, dstrp, fDAPt, fRPt, tftime, aDAPt, aRPt, rpd, tatime

def read_dicom_files(folder_path):
    if not os.path.isdir(folder_path):
        # Άμυνα αν περάσει κατά λάθος αρχείο
        return [], None

    data_total = []
    dicom_files = list_dicom_files(folder_path)
      # Add proper file extensions if needed
    first_file_processed = False
    dose_report_found = False
    for file in dicom_files:
        file_path = os.path.join(folder_path, file)
        dicom_data = pydicom.dcmread(file_path)
        first_file_processed = True
        sop_class_uid = dicom_data.get((0x0008, 0x0016), None)
        pname = ''.join(dicom_data.get('PatientName', 'N/A'))
        pname = pname.replace('^', ' ')
        study_date_str = dicom_data.get('StudyDate', 'N/A')
        if study_date_str != 'N/A':
            content_date = dt.strptime(study_date_str, '%Y%m%d')

        modality = dicom_data.get("Modality", "")
        series_desc = dicom_data.get("SeriesDescription", "")

        if ((sop_class_uid and (
                sop_class_uid.value == "1.2.840.10008.5.1.4.1.1.88.67"  # Standard X-Ray Radiation Dose SR
                or sop_class_uid.value == "1.2.840.10008.5.1.4.1.1.88.68"  # Enhanced Dose SR
                ))
                or (modality == "SR" and "dose" in series_desc.lower())  # vendor SR fallback
                ):

            DAPtotal, RPt, dstrp, fDAPt, fRPt, tftime, aDAPt, aRPt, rpd, tatime = extract_data(dicom_data)
            dose_report_found = True
            
            # Patient Age
            if (0x0010, 0x1010) in dicom_data:
                age = dicom_data[(0x0010, 0x1010)].value
                age = extract_and_format_age(age)
            else:
                # Calculate the age using the patient's birth date and the study date
                birth_date_str = dicom_data[(0x0010, 0x0030)].value  # Patient's Birth Date
                if birth_date_str != 'N/A' and study_date_str != 'N/A':

                    birth_date = dt.strptime(birth_date_str, '%Y%m%d')
                    age = content_date.year - birth_date.year - (
                            (content_date.month, content_date.day) < (birth_date.month, birth_date.day))

                else:
                    age = 'N/A'

            data_total.append({'Patient Name': pname,
                    'Patient ID': dicom_data.get('PatientID', 'N/A'),
                    'Age (years)': age,
                    'Manufacturer': dicom_data.get('Manufacturer', 'N/A'),
                    'Study Description': dicom_data.get('StudyDescription', 'N/A'),
                    'Station Name': dicom_data.get('StationName', 'N/A'),
                    'Institution Name': dicom_data.get('InstitutionName', 'N/A'),
                    'Study Date': study_date_str,
                    'Performing Physician': dicom_data.get('PerformingPhysicianName', 'N/A'),
                    'Dose Area Product Total (Gym²)': DAPtotal[0] if len(DAPtotal) > 0 else 'N/A',
                    'Dose (RP) Total (Gy)': RPt[0] if len(RPt) > 0 else 'N/A',
                    'Fluoro Dose Area Product Total (μGym²)': fDAPt[0] if len(fDAPt) > 0 else 'N/A',
                    'Fluoro Dose (RP) Total (Gy)': fRPt[0] if len(fRPt) > 0 else 'N/A',
                    'Total Fluoro Time (s)': tftime[0] if len(tftime) > 0 else 'N/A',
                    'Acquisition Dose Area Product Total (Gym²)': aDAPt[0] if len(aDAPt) > 0 else 'N/A',
                    'Acquisition Dose (RP) Total (Gy)': aRPt[0] if len(aRPt) > 0 else 'N/A',
                    'Reference Point Definition (cm)': rpd[0] if len(rpd) > 0 else 'N/A',
                    'Total Acquisition Time (s)': tatime[0] if len(tatime) > 0 else 'N/A'})
        else:
            pass

    if not dose_report_found:
        print(f"No dose report found in folder {folder_path}")
    if first_file_processed:
        dftotal = pd.DataFrame(data_total)

        return dftotal, pname
    else:
        return None, None

def list_dicom_files(folder):
    """Επιστρέφει ΜΟΝΟ έγκυρα DICOM αρχεία: .dcm ή χωρίς επέκταση, όχι κρυφά/mac σκουπιδάκια."""
    return [
        entry.path
        for entry in os.scandir(folder)
        if entry.is_file()
        and not entry.name.startswith(".")  # κόβει .DS_Store και ._*
        and (entry.name.lower().endswith(".dcm") or os.path.splitext(entry.name)[1] == "")
    ]

author = os.getlogin()
# open a window
root = tk.Tk()
root.withdraw()
# Select DICOM folder
folder_path = filedialog.askdirectory(title="Select Folder with DICOM DATA")
folder_path = sanitize_path(folder_path)
# Select output folder for the new Excel file, starting from the parent of the DICOM folder
initial_output_dir = os.path.dirname(folder_path)
output_folder = filedialog.askdirectory(title="Select Folder to Save the Excel File", initialdir=initial_output_dir)
output_folder = sanitize_path(output_folder)
# Automatically create an Excel file path
timestamp = time.strftime("%Y-%m-%d_%H%M%S")
output_filename = f"Dose_Report_{timestamp}.xlsx"
output_directory = os.path.join(output_folder, output_filename)

# Reference Value Levels
print("Reference Value Levels")
def get_float(prompt):
    try:
        return float(input(prompt))
    except ValueError:
        print(f"Invalid input for {prompt.strip(':')}.")
        return None
DAP_value = get_float("Dose Area Product (DAP) (Gym²):")
DOSERP_value = get_float("Dose (RP) (Gy):")
AT_value = get_float("Acquisition Time (s):")

files_here = list_dicom_files(folder_path)

if files_here:
    # Χειρισμός περίπτωσης: διάλεξες απευθείας φάκελο ασθενή
    patient_dirs = [folder_path]
else:
    # Συνήθης περίπτωση: διάλεξες root που περιέχει φακέλους ασθενών
    patient_dirs = [
        entry.path for entry in os.scandir(folder_path)
        if entry.is_dir() and not entry.name.startswith(".")
    ]

total = []
coun = 0

for pdir in patient_dirs:
    dftotal, pname = read_dicom_files(pdir)
    
    #--------------------

    if dftotal is not None and not dftotal.empty:
        total.append(dftotal)
        coun += 1
        print(f"Processed file: {coun}")
    else:
        print(f"Skipped folder (no data): {pdir}")
    
if total:
    with pd.ExcelWriter(output_directory, engine='openpyxl') as writer:
        # Συνένωση όλων των DataFrames
        dft = pd.concat(total, axis=0, ignore_index=True)

        # Κλειδιά ομαδοποίησης
        key_cols = ['Patient Name', 'Patient ID']

        # Δηλώνουμε ρητά ποιες στήλες είναι αριθμητικές για άθροιση
        numeric_cols = [
            'Dose Area Product Total (Gym²)',
            'Dose (RP) Total (Gy)',
            'Fluoro Dose Area Product Total (μGym²)',
            'Fluoro Dose (RP) Total (Gy)',
            'Total Fluoro Time (s)',
            'Acquisition Dose Area Product Total (Gym²)',
            'Acquisition Dose (RP) Total (Gy)',
            'Total Acquisition Time (s)',
            # βάλε κι άλλες ΜΟΝΟ αν είσαι βέβαιος ότι είναι αριθμητικές
        ]

        # Προσοχή: ΜΗΝ πειράξεις Patient ID / Study Date (μένουν string)
        # Κάνουμε numeric coercion μόνο στις δηλωμένες numeric_cols
        for col in numeric_cols:
            if col in dft.columns:
                dft[col] = pd.to_numeric(dft[col], errors='coerce')

        # Συνάρτηση για μη-αριθμητικές: πρώτη μη κενή/μη "N/A"/μη "empty"
        def first_nonempty(series):
            for v in series:
                if pd.notna(v) and v not in ('N/A', 'empty', ''):
                    return v
            return 'N/A'

        # Χτίζουμε agg dict
        agg_funcs = {}
        for col in dft.columns:
            if col in key_cols:
                continue
            if col in numeric_cols:
                agg_funcs[col] = 'sum'
            else:
                agg_funcs[col] = first_nonempty

        # Ομαδοποίηση/συγχώνευση σε μία γραμμή ανά (Patient Name, Patient ID)
        dft = dft.groupby(key_cols, as_index=False).agg(agg_funcs)

        # ΜΟΝΟ στις αριθμητικές στήλες: 0 -> NaN -> "empty"
        for col in numeric_cols:
            if col in dft.columns:
                dft[col] = dft[col].replace(0, float('nan'))
        dft = dft.fillna('empty')

        # === ΣTYΛ: wrap + border (όπως το είχες) ===
        sty = (
            dft.style
            .set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            .set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
        )
        sty.to_excel(writer, sheet_name="Accumulated X-Ray Dose Data", index=False)

else:
    print("No valid data to write. Excel file was not created.")

if os.path.exists(output_directory):
    wb = openpyxl.load_workbook(output_directory)
    visible_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == 'visible']
    if not visible_sheets:
        wb.active = wb.sheetnames[0]

    sheet1 = wb["Accumulated X-Ray Dose Data"]
    sheet1.row_dimensions[1].height = 30

    # Map header -> column index
    header_map = {cell.value: cell.column for cell in sheet1[1] if cell.value}

    col_DAP   = header_map.get('Dose Area Product Total (Gym²)')
    col_DOSERP= header_map.get('Dose (RP) Total (Gy)')
    col_AT    = header_map.get('Total Acquisition Time (s)')

    # === ΝΕΟ: εξαναγκάζουμε text format σε Patient ID και Study Date ===
    from openpyxl.styles import Alignment
    alignment_settings = Alignment(wrap_text=True, horizontal='left')

    col_pid = header_map.get('Patient ID')
    col_sdt = header_map.get('Study Date')

    if col_pid:
        for r in range(2, sheet1.max_row + 1):
            c = sheet1.cell(row=r, column=col_pid)
            c.number_format = '@'   # Text
    if col_sdt:
        for r in range(2, sheet1.max_row + 1):
            c = sheet1.cell(row=r, column=col_sdt)
            c.number_format = '@'   # Text

    # === Βάψιμο / σχόλια με βάση thresholds ===
    from openpyxl.styles import PatternFill
    for i in range(2, sheet1.max_row + 1):
        if col_DAP:
            cell = sheet1.cell(row=i, column=col_DAP)
            try:
                val = float(cell.value)
                if val > 0.05:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.comment = Comment(f"Exceeds trigger level: 0.05 Gym²", author)
                elif DAP_value is not None and val > DAP_value:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.comment = Comment(f"Exceeds reference level: {DAP_value} Gym²", author)
            except (TypeError, ValueError):
                pass

        if col_DOSERP:
            cell = sheet1.cell(row=i, column=col_DOSERP)
            try:
                val = float(cell.value)
                if val > 5:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.comment = Comment(f"Exceeds trigger level: 5 Gy", author)
                elif DOSERP_value is not None and val > DOSERP_value:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.comment = Comment(f"Exceeds reference level: {DOSERP_value} Gy", author)
            except (TypeError, ValueError):
                pass

        if col_AT:
            cell = sheet1.cell(row=i, column=col_AT)
            try:
                val = float(cell.value)
                if val > 3600:
                    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    cell.comment = Comment(f"Exceeds trigger level: 1 hour", author)
                elif AT_value is not None and val > AT_value:
                    cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    cell.comment = Comment(f"Exceeds reference value: {AT_value} s", author)
            except (TypeError, ValueError):
                pass

    # Header wrap (κρατάμε τη στοίχιση)
    for col in range(1, sheet1.max_column + 1):
        sheet1.cell(row=1, column=col).alignment = alignment_settings

    # (Προαιρετικό) πλάτος στηλών
    from openpyxl.utils import get_column_letter
    default_width = 15
    for col in range(1, sheet1.max_column + 1):
        sheet1.column_dimensions[get_column_letter(col)].width = default_width

    wb.save(output_directory)
else:
    print("Excel file was not created. No data to load.")



print("\nProcessing done!")