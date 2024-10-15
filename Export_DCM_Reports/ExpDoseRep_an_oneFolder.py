import pydicom
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime as dt
import time

class DicomFile:
    def __init__(self, filename, file_path):
        self.filename = filename
        self.file_path = file_path

def sanitize_path(path):
    return path.strip('"')
def read_hex_to_decimal(dicom_data, tag):
    """Read a hexadecimal value from a DICOM tag and convert it to decimal."""
    if tag in dicom_data:
        values = dicom_data[tag].value
        if values is None:
            raise TypeError(f"Tag {tag} found but has no value (None).")

        if isinstance(values, list):
            decimal_values = []
            for value in values:
                if isinstance(value, bytes):
                    hex_value = value.hex()
                elif isinstance(value, str):
                    hex_value = value
                elif isinstance(value, int):
                    decimal_values.append(value)
                    continue
                else:
                    raise TypeError(f"Unexpected type for tag {tag}: {type(value)}")

                decimal_value = int(hex_value, 16)
                decimal_values.append(decimal_value)
            return decimal_values
        else:
            if isinstance(values, bytes):
                hex_value = values.hex()
            elif isinstance(values, str):
                hex_value = values
            elif isinstance(values, int):
                return values
            else:
                raise TypeError(f"Unexpected type for tag {tag}: {type(values)}")
            decimal_value = int(hex_value, 16)
            return decimal_value
    else:
        return 'N/A'
def extract_and_format_age(age_value):
    """Extract the numeric part of the age and format it."""
    if isinstance(age_value, str):
        numeric_part = ''.join(filter(str.isdigit, age_value))
        if numeric_part:
            return int(numeric_part)
        else:
            raise ValueError(f"No numeric part found in value '{age_value}'")
def sanitize_sheet_name(sheet_name):
    return "".join([c for c in sheet_name if c.isalnum() or c in [' ', '_', '-']]).strip()

def extract_data(dicom_data):
    # ==Totals===
    DAPtotal = []
    RPt = []
    dstrp = []
    fDAPt = []
    fRPt = []
    tftime = []
    aDAPt = []
    aRPt = []
    rpd = []
    tatime = []
    # ===Events==
    event_type = []
    DAP = []
    drp = []
    primang = []
    secang = []
    xrayfiltype = []
    xraymat = []
    thicmax = []
    thicmin = []
    pulse_rate = []
    numb_pulses = []
    irrad_dur = []
    KVP = []
    current = []
    exp_time = []
    pulse_width = []
    exposure = []
    cfield_area = []
    cfield_height = []
    cfield_width = []
    ds_toiso = []
    ds_todet = []

    def search_sequence(sequence):
        global count  # Ensure the 'count' variable is accessible globally
        count = False  # Initialize 'count' at the start of the sequence search

        for item in sequence:
            # Check if the item has the Concept Name Code Sequence (0040,A043)
            if (0x0040, 0xA043) in item:
                concept_name_code_sequence = item[(0x0040, 0xA043)].value
                code_meaning = concept_name_code_sequence[0].get((0x0008, 0x0104), None)
                if code_meaning:
                    code_meaning_value = code_meaning.value
                    if code_meaning_value == 'Dose Area Product Total':
                        extract_numeric_value(item, (0x0040, 0xA300), DAPtotal)
                    elif code_meaning_value == 'Dose (RP) Total':
                        extract_numeric_value(item, (0x0040, 0xA300), RPt)
                    elif code_meaning_value == 'Distance Source to Reference Point':
                        extract_numeric_value(item, (0x0040, 0xA300), dstrp)
                    elif code_meaning_value == 'Fluoro Dose Area Product Total':
                        extract_numeric_value(item, (0x0040, 0xA300), fDAPt)
                    elif code_meaning_value == 'Fluoro Dose (RP) Total':
                        extract_numeric_value(item, (0x0040, 0xA300), fRPt)
                    elif code_meaning_value == 'Total Fluoro Time':
                        extract_numeric_value(item, (0x0040, 0xA300), tftime)
                    elif code_meaning_value == 'Acquisition Dose Area Product Total':
                        extract_numeric_value(item, (0x0040, 0xA300), aDAPt)
                    elif code_meaning_value == 'Acquisition Dose (RP) Total':
                        extract_numeric_value(item, (0x0040, 0xA300), aRPt, allow_empty=True)
                    elif code_meaning_value == 'Irradiation Event Type':
                        extract_concept_code(item, (0x0040, 0xA168), event_type)
                    elif code_meaning_value == 'Reference Point Definition':
                        extract_concept_code(item, (0x0040, 0xA168), rpd)
                    elif code_meaning_value == 'Total Acquisition Time':
                        extract_numeric_value(item, (0x0040, 0xA300), tatime)
                    elif code_meaning_value == 'Dose Area Product':
                        extract_numeric_value(item, (0x0040, 0xA300), DAP)
                    elif code_meaning_value == 'Dose (RP)':
                        extract_numeric_value(item, (0x0040, 0xA300), drp)
                    elif code_meaning_value == 'Positioner Primary Angle':
                        extract_numeric_value(item, (0x0040, 0xA300), primang)
                    elif code_meaning_value == 'Positioner Secondary Angle':
                        extract_numeric_value(item, (0x0040, 0xA300), secang)
                    elif code_meaning_value == 'X-Ray Filter Type':
                        extract_concept_code(item, (0x0040, 0xA168), xrayfiltype, filter_check=True)
                    elif code_meaning_value == 'X-Ray Filter Material':
                        extract_concept_code(item, (0x0040, 0xA168), xraymat)
                    elif code_meaning_value == 'X-Ray Filter Thickness Maximum':
                        extract_numeric_value(item, (0x0040, 0xA300), thicmax, allow_empty=True)
                    elif code_meaning_value == 'X-Ray Filter Thickness Minimum':
                        extract_numeric_value(item, (0x0040, 0xA300), thicmin, allow_empty=True)
                    elif code_meaning_value == 'Pulse Rate':
                        extract_numeric_value(item, (0x0040, 0xA300), pulse_rate)
                    elif code_meaning_value == 'Number of Pulses':
                        extract_numeric_value(item, (0x0040, 0xA300), numb_pulses)
                    elif code_meaning_value == 'Irradiation Duration':
                        if not count:
                            extract_numeric_value(item, (0x0040, 0xA300), irrad_dur)
                            count = True
                    elif code_meaning_value == 'KVP':
                        extract_numeric_value(item, (0x0040, 0xA300), KVP)
                    elif code_meaning_value == 'X-Ray Tube Current':
                        extract_numeric_value(item, (0x0040, 0xA300), current)
                    elif code_meaning_value == 'Exposure Time':
                        extract_numeric_value(item, (0x0040, 0xA300), exp_time)
                    elif code_meaning_value == 'Pulse Width':
                        extract_numeric_value(item, (0x0040, 0xA300), pulse_width)
                    elif code_meaning_value == 'Exposure':
                        extract_numeric_value(item, (0x0040, 0xA300), exposure)
                    elif code_meaning_value == 'Collimated Field Area':
                        extract_numeric_value(item, (0x0040, 0xA300), cfield_area)
                    elif code_meaning_value == 'Collimated Field Height':
                        extract_numeric_value(item, (0x0040, 0xA300), cfield_height)
                    elif code_meaning_value == 'Collimated Field Width':
                        extract_numeric_value(item, (0x0040, 0xA300), cfield_width)
                    elif code_meaning_value == 'Distance Source to Isocenter':
                        extract_numeric_value(item, (0x0040, 0xA300), ds_toiso)
                    elif code_meaning_value == 'Distance Source to Detector':
                        extract_numeric_value(item, (0x0040, 0xA300), ds_todet)

            # Recursively search within nested sequences
            if (0x0040, 0xA730) in item:
                search_sequence(item[(0x0040, 0xA730)].value)

    # Function to extract numeric value from Measured Value Sequence
    def extract_numeric_value(item, tag, target_list, allow_empty=False):
        if tag in item:
            measured_value_sequence = item[tag].value
            if measured_value_sequence and (0x0040, 0xA30A) in measured_value_sequence[0]:
                numeric_value = measured_value_sequence[0][(0x0040, 0xA30A)].value
                target_list.append(float(numeric_value))
            elif allow_empty:
                target_list.append('empty')
            else:
                target_list.append("N/A")

    # Function to extract code value from Concept Code Sequence
    def extract_concept_code(item, tag, target_list, filter_check=False):
        if tag in item:
            concept_code_sequence = item[tag].value
            if concept_code_sequence and (0x0008, 0x0104) in concept_code_sequence[0]:
                code_value = concept_code_sequence[0][(0x0008, 0x0104)].value
                target_list.append(code_value)
                if filter_check and code_value == 'No filter':
                    xraymat.append('N/A')
                    thicmax.append('N/A')
                    thicmin.append('N/A')
            else:
                target_list.append('N/A')

    # Start searching from the main sequence
    if (0x0040, 0xA730) in dicom_data:
        search_sequence(dicom_data[(0x0040, 0xA730)].value)
    filttype1 = []
    mat1 = []
    tmin1 = []
    tmax1 = []
    filttype2 = []
    mat2 = []
    tmin2 = []
    tmax2 = []
    for i in range(1, len(xraymat), 2):
        if str(xraymat[i - 1]) != str(xraymat[i]):
            filttype1.append(xrayfiltype[i])
            filttype2.append(xrayfiltype[i - 1])
            mat1.append(xraymat[i])
            mat2.append(xraymat[i - 1])
            if thicmin:
                tmin1.append(thicmin[i])
                tmin2.append(thicmin[i - 1])
            if thicmax:
                tmax1.append(thicmax[i])
                tmax2.append(thicmax[i - 1])
    if mat1 and mat2:
        max_len = max(len(lst) for lst in
                      [DAP, drp, primang, secang, filttype1, mat1, tmin1, tmax1, filttype2, mat2, tmin2, tmax2,
                       pulse_rate, numb_pulses, irrad_dur, KVP, current, exp_time, pulse_width, exposure, cfield_area,
                       cfield_height, cfield_width, ds_toiso, ds_todet])
    else:
        max_len = max(len(lst) for lst in
                      [DAP, drp, primang, secang, xrayfiltype, xraymat, thicmax, thicmin, pulse_rate, numb_pulses,
                       irrad_dur, KVP, current, exp_time, pulse_width, exposure, cfield_area, cfield_height,
                       cfield_width, ds_toiso, ds_todet])
    # Function to pad lists with 'empty' so they all have the same length
    def pad_list(lst, max_len):
        return lst + ['N/A'] * (max_len - len(lst))

    # Pad each list
    DAP = pad_list(DAP, max_len)
    drp = pad_list(drp, max_len)
    primang = pad_list(primang, max_len)
    secang = pad_list(secang, max_len)
    xrayfiltype = pad_list(xrayfiltype, max_len)
    xraymat = pad_list(xraymat, max_len)
    thicmax = pad_list(thicmax, max_len)
    thicmin = pad_list(thicmin, max_len)
    filttype1 = pad_list(filttype1, max_len)
    mat1 = pad_list(mat1, max_len)
    tmax1 = pad_list(tmax1, max_len)
    tmin1 = pad_list(tmin1, max_len)
    filttype2 = pad_list(filttype2, max_len)
    mat2 = pad_list(mat2, max_len)
    tmax2 = pad_list(tmax2, max_len)
    tmin2 = pad_list(tmin2, max_len)
    pulse_rate = pad_list(pulse_rate, max_len)
    numb_pulses = pad_list(numb_pulses, max_len)
    irrad_dur = pad_list(irrad_dur, max_len)
    KVP = pad_list(KVP, max_len)
    current = pad_list(current, max_len)
    exp_time = pad_list(exp_time, max_len)
    pulse_width = pad_list(pulse_width, max_len)
    exposure = pad_list(exposure, max_len)
    cfield_area = pad_list(cfield_area, max_len)
    cfield_height = pad_list(cfield_height, max_len)
    cfield_width = pad_list(cfield_width, max_len)
    ds_toiso = pad_list(ds_toiso, max_len)
    ds_todet = pad_list(ds_todet, max_len)

    return DAPtotal, RPt, dstrp, fDAPt, fRPt, tftime, aDAPt, aRPt, rpd, tatime, DAP, drp, primang, secang, \
        xrayfiltype,  xraymat, thicmax, thicmin,filttype1,mat1,tmin1,tmax1,filttype2,mat2,tmin2,tmax2, pulse_rate, \
        numb_pulses, irrad_dur, KVP, current, exp_time, \
        pulse_width, exposure, cfield_area, cfield_height, cfield_width, ds_toiso, ds_todet,event_type, max_len

def read_dicom_files(folder_path, coun, rsname):
    data_total = []
    info_dict = {}
    data_all_st = []
    data_all_fl = []
    #dicom_files = [file for file in os.listdir(folder_path) if file.endswith('.dcm') or file.endswith('')]  # Add proper file extensions if needed
    first_file_processed = False
    file_counts = 0
    series_counts = 0
    events =0
    dose_report_found = False
    """for file in dicom_files:
        file_path = os.path.join(folder_path, file)"""
    file_counts += 1
    dicom_data = pydicom.dcmread(folder_path)

    if (0x0020, 0x0011) in dicom_data and (0x0008, 0x103E) in dicom_data:
        exam_protocol_sr = dicom_data[(0x0008, 0x103E)].value

        if str(exam_protocol_sr) == str(rsname) :
            DAPtotal, RPt, dstrp, fDAPt, fRPt, tftime, aDAPt, aRPt, rpd, tatime, DAP, drp, primang, secang, \
            xrayfiltype, xraymat, thicmax, thicmin, filttype1, mat1, tmin1, tmax1, filttype2, mat2, tmin2, tmax2, \
            pulse_rate, numb_pulses, irrad_dur, KVP, current, exp_time, pulse_width, exposure, cfield_area, cfield_height, cfield_width, \
            ds_toiso, ds_todet, event_type, events = extract_data(dicom_data)

            dose_report_found = True

            ev_st = 0
            ev_fl = 0
            for i in range(0, events):
                if event_type[i] == 'Stationary Acquisition':
                    ev_st += 1
                    data_all_st.append({"Dose Area Product (Gym²)": DAP[i], "Dose (RP) (Gy)": drp[i],
                                        'Positioner Primary Angle (deg)': primang[i],
                                        'Positioner Secondary Angle (deg)': secang[i],
                                        'X-Ray Filter Type': xrayfiltype[i],
                                        'X-Ray Filter Thickness Material': xraymat[i],
                                        'X-Ray Filter Thickness Maximum (mmCu)': thicmax[i],
                                        'X-Ray Filter Thickness Minimum (mmCu)': thicmin[i],
                                        "Pulse Rate (pulse/s)": pulse_rate[i],
                                        "Irradiation Duration (s)": irrad_dur[i], 'KVP': KVP[i],
                                        'X-Ray Tube Current (mA)': current[i], 'Exposure Time (ms)': exp_time[i],
                                        'Pulse Width (ms)': pulse_width[i], 'Exposure (uA.s)': exposure[i],
                                        'Collimated Field Area (m²)': cfield_area[i],
                                        'Collimated Field Height (mm)': cfield_height[i],
                                        'Collimated Field Width (mm)': cfield_width[i],
                                        'Distance Source to Isocenter (mm)': ds_toiso[i],
                                        'Distance Source to Detector (mm)': ds_todet[i]})
                elif event_type[i] == 'Fluoroscopy':
                    ev_fl += 1
                    data_all_fl.append({"Dose Area Product (Gym²)": DAP[i], "Dose (RP) (Gy)": drp[i],
                                        'Positioner Primary Angle (deg)': primang[i],
                                        'Positioner Secondary Angle (deg)': secang[i],
                                        'X-Ray Filter Type': xrayfiltype[i],
                                        'X-Ray Filter Thickness Material': xraymat[i],
                                        'X-Ray Filter Thickness Maximum (mmCu)': thicmax[i],
                                        'X-Ray Filter Thickness Minimum (mmCu)': thicmin[i],
                                        "Pulse Rate (pulse/s)": pulse_rate[i],
                                        "Irradiation Duration (s)": irrad_dur[i], 'KVP': KVP[i],
                                        'X-Ray Tube Current (mA)': current[i], 'Exposure Time (ms)': exp_time[i],
                                        'Pulse Width (ms)': pulse_width[i], 'Exposure (uA.s)': exposure[i],
                                        'Collimated Field Area (m²)': cfield_area[i],
                                        'Collimated Field Height (mm)': cfield_height[i],
                                        'Collimated Field Width (mm)': cfield_width[i],
                                        'Distance Source to Isocenter (mm)': ds_toiso[i],
                                        'Distance Source to Detector (mm)': ds_todet[i]})
                else:
                    data_all_fl.append({"Dose Area Product (Gym²)": DAP[i], "Dose (RP) (Gy)": drp[i],
                                        'Positioner Primary Angle (deg)': primang[i],
                                        'Positioner Secondary Angle (deg)': secang[i],
                                        'X-Ray Filter Type': xrayfiltype[i],
                                        'X-Ray Filter Thickness Material': xraymat[i],
                                        'X-Ray Filter Thickness Maximum (mmCu)': thicmax[i],
                                        'X-Ray Filter Thickness Minimum (mmCu)': thicmin[i],
                                        "Pulse Rate (pulse/s)": pulse_rate[i],
                                        "Irradiation Duration (s)": irrad_dur[i], 'KVP': KVP[i],
                                        'X-Ray Tube Current (mA)': current[i], 'Exposure Time (ms)': exp_time[i],
                                        'Pulse Width (ms)': pulse_width[i], 'Exposure (uA.s)': exposure[i],
                                        'Collimated Field Area (m²)': cfield_area[i],
                                        'Collimated Field Height (mm)': cfield_height[i],
                                        'Collimated Field Width (mm)': cfield_width[i],
                                        'Distance Source to Isocenter (mm)': ds_toiso[i],
                                        'Distance Source to Detector (mm)': ds_todet[i]})
            if len(mat1) > 0 and len(mat2) > 0 and mat1[0] != 'N/A' and mat2[0] != 'N/A':
                data_all1 = []
                data_all2 = []
                if data_all_st:
                    for idx, item in enumerate(data_all_st):
                        updated_item = {}
                        for key, value in item.items():
                            if key == 'X-Ray Filter Type':
                                updated_item['X-Ray Filter Type 1'] = filttype1[idx]
                                updated_item['X-Ray Filter Type 2'] = filttype2[idx]
                            elif key == 'X-Ray Filter Thickness Material':
                                updated_item['X-Ray Filter Thickness Material 1'] = mat1[idx]
                                updated_item['X-Ray Filter Thickness Material 2'] = mat2[idx]
                            elif key == 'X-Ray Filter Thickness Maximum (mmCu)':
                                updated_item['X-Ray Filter Thickness Maximum (mmCu) 1'] = tmax1[idx]
                                updated_item['X-Ray Filter Thickness Maximum (mmCu) 2'] = tmax2[idx]
                            elif key == 'X-Ray Filter Thickness Minimum (mmCu)':
                                updated_item['X-Ray Filter Thickness Minimum (mmCu) 1'] = tmin1[idx]
                                updated_item['X-Ray Filter Thickness Minimum (mmCu) 2'] = tmin2[idx]
                            else:
                                updated_item[key] = value
                        data_all1.append(updated_item)
                    data_all_st = data_all1
                if data_all_fl:
                    for idx, item in enumerate(data_all_fl):
                        updated_item = {}
                        for key, value in item.items():
                            if key == 'X-Ray Filter Type':
                                updated_item['X-Ray Filter Type 1'] = filttype1[idx]
                                updated_item['X-Ray Filter Type 2'] = filttype2[idx]
                            elif key == 'X-Ray Filter Thickness Material':
                                updated_item['X-Ray Filter Thickness Material 1'] = mat1[idx]
                                updated_item['X-Ray Filter Thickness Material 2'] = mat2[idx]
                            elif key == 'X-Ray Filter Thickness Maximum (mmCu)':
                                updated_item['X-Ray Filter Thickness Maximum (mmCu) 1'] = tmax1[idx]
                                updated_item['X-Ray Filter Thickness Maximum (mmCu) 2'] = tmax2[idx]
                            elif key == 'X-Ray Filter Thickness Minimum (mmCu)':
                                updated_item['X-Ray Filter Thickness Minimum (mmCu) 1'] = tmin1[idx]
                                updated_item['X-Ray Filter Thickness Minimum (mmCu) 2'] = tmin2[idx]
                            else:
                                updated_item[key] = value
                        data_all2.append(updated_item)
                    data_all_fl = data_all2
            study_date_str = dicom_data.get('StudyDate', 'N/A')
            if study_date_str != 'N/A':
                content_date = dt.strptime(study_date_str, '%Y%m%d')
                date_str = content_date.strftime('%Y-%m-%d')
                info_dict['Content Date'] = date_str
            physician = f"Physician {coun}"
            data_total.append(
                {"Patient ID": dicom_data.get('PatientID', 'N/A'), "Manufacturer": dicom_data.get('Manufacturer', 'N/A'),
                    "Content Date": date_str, 'Performing Physician': physician,
                    'Dose Area Product Total (μGym²)': DAPtotal[0] if len(DAPtotal) > 0 else 'N/A',
                    'Dose (RP) Total (mGy)': RPt[0] if len(RPt) > 0 else 'N/A',
                    'Fluoro Dose Area Product Total (μGym²)': fDAPt[0] if len(fDAPt) > 0 else 'N/A',
                    "Fluoro Dose (RP) Total (Gy)": fRPt[0] if len(fRPt) > 0 else 'N/A',
                    "Total Fluoro Time (s)": tftime[0] if len(tftime) > 0 else 'N/A',
                    "Acquisition Dose Area Product Total (Gym²)": aDAPt[0] if len(aDAPt) > 0 else 'N/A',
                    "Acquisition Dose (RP) Total (Gy)": aRPt[0] if len(aRPt) > 0 else 'N/A',
                    "Reference Point Definition (cm)": rpd[0] if len(rpd) > 0 else 'N/A',
                    "Total Acquisition Time (s)": tatime[0] if len(tatime) > 0 else 'N/A'})


        else:
            pass
        series_counts += 1
        if not first_file_processed:
            physician = f"Physician {coun}"
            pname = f'Patient {coun}'
            info_dict = {'Patient Name': pname, 'Patient ID': dicom_data.get('PatientID', 'N/A'),
                'Content Date': dicom_data.get('StudyDate', 'N/A'),
                'Performing Physician': physician}
            psex = dicom_data[(0x0010, 0x0040)].value

            if psex == 'F':
                psex = 'Female'
            else:
                psex = 'Male'
            first_file_processed = True
            study_date_str = dicom_data.get('StudyDate', 'N/A')
            if study_date_str != 'N/A':
                content_date = dt.strptime(study_date_str, '%Y%m%d')
                date_str = content_date.strftime('%Y-%m-%d')
                info_dict['Content Date'] = date_str

            content_time = dicom_data.get('ContentTime','N/A')
            if content_time != 'N/A':
                # Convert the Content Time to HH:MM:SS format
                time_str = content_time[:2] + ':' + content_time[2:4] + ':' + content_time[4:6]
            else:
                print("ContentTime not available")

            if (0x0010,0x1010) in dicom_data:
                age = dicom_data[(0x0010,0x1010)].value
                age = extract_and_format_age(age)
            else:
                # Calculate the age using the patient's birth date and the study date
                birth_date_str = dicom_data[(0x0010, 0x0030)].value  # Patient's Birth Date
                if birth_date_str != 'N/A' and study_date_str != 'N/A':

                    birth_date = dt.strptime(birth_date_str, '%Y%m%d')
                    age = content_date.year - birth_date.year - (
                                (content_date.month, content_date.day) < (birth_date.month, birth_date.day))
                    age_str = f"{age}Y"  # Format age as years (e.g., '45Y')
                else:
                    age_str = 'N/A'

            infoplus = {'Patient Name': pname,'Patient ID': dicom_data.get('PatientID', 'N/A'),
            "Gender": psex,'Age (years)':age,'Study Description':dicom_data.get('StudyDescription', 'N/A'),
            "Manufacturer":dicom_data.get("Manufacturer", 'N/A'),'Content Date': date_str,'Content Time':time_str,
            'Performing Physician': physician}
    if not dose_report_found:
        print(f"No dose report found in folder {folder_path}")

    if first_file_processed:
        df_fl = None
        df_st = None
        dfper_fl = None
        dfper_st = None
        if data_all_fl:
            info_dict['Number of Events'] = ev_fl
            df_fl = pd.DataFrame(data_all_fl)
            df_fl = df_fl.fillna('Ν/Α')
            df_fl = df_fl.rename_axis(f'Irradiation Event X-Ray Data of {pname}')
            for i in range(0, ev_fl + 1):
                df_fl = df_fl.rename(index={i: f"Event {i + 1}"})
            dfper_fl = pd.DataFrame([info_dict]).T
            dfper_fl.columns = ['']
            dfper_fl.index.name = 'Patient Info'
        if data_all_st:
            info_dict['Number of Events'] = ev_st
            df_st = pd.DataFrame(data_all_st)
            df_st = df_st.fillna('Ν/Α')
            df_st = df_st.rename_axis(f'Irradiation Event X-Ray Data of {pname}')
            for i in range(0, ev_st + 1):
                df_st = df_st.rename(index={i: f"Event {i + 1}"})
            dfper_st = pd.DataFrame([info_dict]).T
            dfper_st.columns = ['']
            dfper_st.index.name = 'Patient Info'
        dftotal = pd.DataFrame(data_total)
        dfinfo = pd.DataFrame([infoplus])
        dfinfo = dfinfo.rename_axis('Patient Name')
        dfinfo = dfinfo.fillna('N/A')

        return df_fl, df_st, dftotal, dfper_fl, dfper_st, pname, dfinfo, event_type
    else:
        print("No valid DICOM files processed.")
        return None, None, None, None, None, None, None, None

def auto_adjust_column_widths(sheet, start_row_df, max_row, max_col):
    column_widths = {}

    for col in range(1, max_col + 1):
        max_length = 0
        for row in range(start_row_df + 1, max_row + 2):
            cell = sheet.cell(row=row, column=col)
            if cell.value:
                # Calculate the length of the cell text with respect to wrapping
                cell_value = str(cell.value)
                lines = cell_value.split('\n')
                max_line_length = max(len(line) for line in lines)
                max_length = max(max_length, max_line_length)

        # Set the width to fit two lines of text with some padding
        column_widths[col] = (max_length / 2) + 5  # Adjust the padding value as needed

    for col, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width

folder_path = sanitize_path(input(r'Write the path of the folder with DICOM DATA: '))
output_directory = sanitize_path(input(r"Write the path of the excel file: "))
rsname = input (r'Write the series description of Radiation Dose Report: ')
start_time = time.time()

dicom_files = [file for file in os.listdir(folder_path) if file.endswith('') or file.endswith('.dcm')]  # Add proper file extensions if needed

coun = 0
d = {}
total = []
info = []
nfl = 0
nst = 0

with pd.ExcelWriter(output_directory, engine='openpyxl') as writer:
    for file in dicom_files:
        file_path = os.path.join(folder_path, file)

        df_fl, df_st, dftotal, dfper_fl, dfper_st, pname, dfinfo, event_type = read_dicom_files(file_path, coun, rsname)
        coun += 1
        print(f"Process file:{coun}")
        if df_fl is not None and dfper_fl is not None:
            nfl += 1
            df_fl.replace(0, "empty", inplace=True)
            for i in range(len(event_type)):
                if event_type [i] == 'Fluoroscopy':
                    sheet_name_fl = "Fluoro-" + str(pname) if pname else "Sheet1"
                    break
                else:
                    sheet_name_fl = str(pname) if pname else "Sheet1"
            sheet_name_fl = sanitize_sheet_name(sheet_name_fl)
            start_row_df_fl = len(dfper_fl) + 2
            dfper_styled = dfper_fl.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            df_styled = df_fl.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            df1_fl = df_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
            df2_fl = dfper_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})

            df2_fl.to_excel(writer, sheet_name=sheet_name_fl, header=False)
            df1_fl.to_excel(writer, sheet_name=sheet_name_fl, startrow=start_row_df_fl)

            d[f"sheetfl {nfl - 1}"] = sheet_name_fl
            d[f"max_rowfl {nfl - 1}"] = max(start_row_df_fl + len(df_fl), len(dfper_fl))
            d[f"max_colfl {nfl - 1}"] = len(df_fl.columns) + 1
        if df_st is not None and dfper_st is not None:
            nst += 1
            df_st.replace(0, "empty", inplace=True)
            sheet_name_st = "DSA-" + str(pname) if pname else "Sheet1"
            sheet_name_st = sanitize_sheet_name(sheet_name_st)
            start_row_df_st = len(dfper_st) + 2
            dfper_styled = dfper_st.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            df_styled = df_st.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
            df1_st = df_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
            df2_st = dfper_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})

            df2_st.to_excel(writer, sheet_name=sheet_name_st, header=False)
            df1_st.to_excel(writer, sheet_name=sheet_name_st, startrow=start_row_df_st)

            d[f"sheetst {nst - 1}"] = sheet_name_st
            d[f"max_rowst {nst - 1}"] = max(start_row_df_st + len(df_st), len(dfper_st))
            d[f"max_colst {nst - 1}"] = len(df_st.columns) + 1
        total.append(dftotal)
        info.append(dfinfo)
    if total:
        # Concatenate all the DataFrames in the total list
        dft = pd.concat(total, axis=0)
        dfi = pd.concat(info, axis = 0)
        dft.replace(0, "empty", inplace=True)
        dft.columns = pd.Index(
            [f'{col}_{i}' if dft.columns.duplicated()[i] else col for i, col in enumerate(dft.columns)])
        dft.index = pd.Index([f'{idx}_{i}' if dft.index.duplicated()[i] else idx for i, idx in enumerate(dft.index)])
        dfi.columns = pd.Index(
            [f'{col}_{i}' if dfi.columns.duplicated()[i] else col for i, col in enumerate(dfi.columns)])
        dfi.index = pd.Index([f'{idx}_{i}' if dfi.index.duplicated()[i] else idx for i, idx in enumerate(dfi.index)])
        # Apply styling if the columns and index are unique
        dfstpat = dft.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
        dft = dfstpat.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
        dfstin = dfi.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
        dfi = dfstin.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
        # Save the styled DataFrame to Excel
        dft.to_excel(writer, sheet_name=f"Accumulated X-Ray Dose Data", index=False)
        dfi.to_excel(writer, sheet_name=f"Basic Study Indormation", index=False)

wb = openpyxl.load_workbook(output_directory)


if nst != 0:
    for i in range(0, nst):
        sheet = wb[d[f"sheetst {i}"]]
        max_row = d[f"max_rowst {i}"]
        max_col = d[f"max_colst {i}"]
        if not df_st.empty:
            start_row_df = start_row_df_st
        else:
            start_row_df = start_row_df_fl
        sheet.row_dimensions[start_row_df + 1].height = 31
        for row in range(start_row_df + 1, max_row + 2):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')
        auto_adjust_column_widths(sheet, start_row_df, max_row, max_col)
if nfl != 0:
    for i in range(0, nfl):
        sheet = wb[d[f"sheetfl {i}"]]
        max_row = d[f"max_rowfl {i}"]
        max_col = d[f"max_colfl {i}"]
        if not df_st.empty:
            start_row_df = start_row_df_st
        else:
            start_row_df = start_row_df_fl
        sheet.row_dimensions[start_row_df + 1].height = 31
        for row in range(start_row_df + 1, max_row + 2):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')
        auto_adjust_column_widths(sheet, start_row_df, max_row, max_col)
# Ensure at least one sheet is visible
visible_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == 'visible']
if not visible_sheets:
    wb.active = wb.sheetnames[0]

sheet1 = wb["Accumulated X-Ray Dose Data"]
sheet2 = wb["Basic Study Indormation"]
sheet1.row_dimensions[1].height = 31
sheet2.row_dimensions[1].height = 31
alignment_settings = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')
for col in range(1, 14):
    cell1 = sheet1.cell(row=1, column=col)
    cell1.alignment = alignment_settings
    cell2 = sheet2.cell(row=1, column=col)
    cell2.alignment = alignment_settings
column_widths = [12, 13, 15.5, 11, 18, 13, 11, 19, 14, 15, 14,17,15]
for i, width in enumerate(column_widths, start=1):
    column_letter = get_column_letter(i)
    sheet1.column_dimensions[column_letter].width = width
column_widths = [11.1, 12.8, 7, 10, 17, 12.4, 12, 11.9, 14.5]
for i, width in enumerate(column_widths, start=1):
    column_letter = get_column_letter(i)
    sheet2.column_dimensions[column_letter].width = width
if nfl != 0 and nst != 0:
    wb.move_sheet(sheet1,offset=-(nfl+nst))
    wb.move_sheet(sheet2, offset=-(nfl+nst)-1)
else:
    wb.move_sheet(sheet1, offset=-coun)
    wb.move_sheet(sheet2, offset=-coun - 1)
wb.save(output_directory)
end_time = time.time()
elapsed_time = end_time - start_time
print(f"Total processing time for {coun} patient folders: {elapsed_time:.2f} seconds")