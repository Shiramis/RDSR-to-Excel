import pydicom
import os
import pandas as pd
import openpyxl
from datetime import datetime as dt
import time

class dicomFilep:
    def __init__(self, filename, file_path):
        self.filename = filename
        self.file_path = file_path

def sanitize_path(path):
    return path.strip('"')

def read_hex_to_decimal(dicom_data, tag):
    """Read a hexadecimal value from a DICOM tag and convert it to decimal."""
    if tag in dicom_data:
        values = dicom_data[tag].value
        if isinstance(values, list):
            decimal_values = []
            for value in values:
                if isinstance(value, bytes):
                    hex_value = value.hex()
                elif isinstance(value, str):
                    hex_value = value
                elif isinstance(value, int):
                    # If the value is already an integer, no conversion is needed
                    decimal_values.append(value)
                    continue
                else:
                    raise TypeError(f"Unexpected type for tag {tag}: {type(value)}")

                decimal_value = int(hex_value, 16)
                decimal_values.append(decimal_value)
            return decimal_values
        else:
            if isinstance(values, bytes):
                hex_value = values.hex()
            elif isinstance(values, str):
                hex_value = values
            elif isinstance(values, int):
                # If the value is already an integer, no conversion is needed
                return values
            else:
                raise TypeError(f"Unexpected type for tag {tag}: {type(values)}")

            decimal_value = int(hex_value, 16)
            return decimal_value
    return 'N/A'

def read_dicom_files(folder_path):
    data = []
    info_dict = {}

    dicom_files = [file for file in os.listdir(folder_path) if file.endswith('')]
    first_file_processed = False
    file_counts = 0
    series_counts = 0
    sd ={}
    for file in dicom_files:

        file_path = os.path.join(folder_path, file)
        file_counts += 1
        dicom_data = pydicom.dcmread(file_path)
        kvp = dicom_data.get('KVP', 'N/A')
        sd["series {0}".format(str(file_counts))] = dicom_data.get('SeriesNumber','N/A')
        if kvp == 'N/A' or file_counts > 1 and sd["series "+ str(file_counts)] == sd["series " + str(file_counts-1)]:
            continue
        filter_thickness = read_hex_to_decimal(dicom_data, (0x0021, 0x100A))
        acquisition_dose = read_hex_to_decimal(dicom_data, (0x0021, 0x1007))
        acquisition_time = read_hex_to_decimal(dicom_data, (0x0021, 0x101F))
        series_counts += 1
        if not first_file_processed:  # Check if the first file has been processed
            pname = ''.join(dicom_data.get('PatientName', 'N/A'))
            info_dict = {'Patient Name': pname,
                'Patient ID': ''.join(dicom_data.get('PatientID', 'N/A')),
                'Content Date': ''.join(dicom_data.get('StudyDate', 'N/A')),
                'Performing Physician': ''.join(dicom_data.get('PerformingPhysicianName', 'N/A'))}
            first_file_processed = True
            study_date_str = dicom_data.get('StudyDate', 'N/A')
            content_date = dt.strptime(study_date_str, '%Y%m%d') if study_date_str != 'N/A' else 'N/A'
            date_str = content_date.strftime('%Y-%m-%d')
            info_dict['Content Date'] = date_str

        positioner_primary_angle = dicom_data.get('PositionerPrimaryAngle', 'N/A')
        try:
            prim = float(positioner_primary_angle)
            if prim >= 0:
                pproj = "RAO"
            else:
                pproj = 'LAO'
        except ValueError:
            prim = 'N/A'

        positioner_secondary_angle = dicom_data.get('PositionerSecondaryAngle', 'N/A')
        try:
            sec = float(positioner_secondary_angle)
            if sec >= 0:
                sproj = 'CRA'
            else:
                sproj = 'CAU'
        except ValueError:
            sec = 'N/A'
        DAP = dicom_data.get('ImageAndFluoroscopyAreaDoseProduct', 'N/A')
        RP = acquisition_dose
        if DAP != 'N/A':
            DAP = DAP*10
        if RP != 'N/A':
            RP = RP/100
        data.append({'Series Number': sd["series "+ str(file_counts)],
            'KVP': kvp,  # KVP tag
            'X-ray Tube Current (mA)': dicom_data.get('XRayTubeCurrent', 'N/A'),  # X-Ray Tube Current tag
            'Pulse Width (ms)': dicom_data.get('AveragePulseWidth','N/A'), # Pulse width
            'Extra X-ray Filter Thickness (mmCu)': filter_thickness,
            'Field Width (diagonal) (mm)': dicom_data.get('IntensifierSize','N/A'),#'Exposure Time': dicom_data.get('ExposureTime','N/A'),
            'Acquisition Time (s)': acquisition_time ,
            'Frame Rate': dicom_data.get('CineRate','N/A'),
            'Acquisition Dose Area Product (μGym²)': DAP,
            'Acquisition Dose (RP) (mGy)': RP,
            'Primary Angle': prim,
            'Primary Projection': pproj,
            'Secondary Angle': sec,
            'Secondary Projection': sproj,
            'Number of Frames': dicom_data.get('NumberOfFrames','N/A')
        })

    info_dict['Number of Series'] = series_counts
    df = pd.DataFrame(data)
    dfper = pd.DataFrame([info_dict])
    dfper = dfper.T  # Transpose DataFrame to have one column
    dfper.columns = ['']
    dfper.index.name = 'Patient Info'
    return df ,dfper, pname

f = input(r'Write the path of the folder with DICOM DATA: ')
folder_path = sanitize_path(f)
ot = input(r"Write the path of the excel file: ")
output_directory = sanitize_path(ot)

start_time = time.time()

dicom_files = []
for filename in os.listdir(folder_path):
    if filename.endswith(''):
        file_path = os.path.join(folder_path, filename)
        dicom_file = dicomFilep(filename, file_path)
        dicom_files.append(dicom_file)

coun = 0
d = {}
with pd.ExcelWriter(output_directory, engine='openpyxl') as writer:
    for dicom_file in dicom_files:
        coun += 1
        df, dfper, name = read_dicom_files(f"{dicom_file.file_path}")
        df.fillna('N/A', inplace=True)

        sheet_name = f"Patient {coun}" # name
        dfper.iat [0,0] = f"Patient {coun}"
        dfper.iat[1,0] = f"ID {coun}"
        dfper.iat[3,0] =f"Physician {coun}"

        start_row_df =len(dfper) + 2
        # Apply style to dataframes
        dfper_styled = dfper.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
        df_styled = df.style.set_properties(**{'text-align': 'left', 'white-space': 'wrap'})
        df1 = df_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
        df2 = dfper_styled.set_properties(**{'border': '1px solid black', 'border-collapse': 'collapse'})
        print(df1)
        df2.to_excel(writer, sheet_name=sheet_name, header=False)
        df1.to_excel(writer, sheet_name=sheet_name, startrow=start_row_df, index=False)

        d["sheet {0}".format(str(coun))] = sheet_name
        d["max_row {0}".format(str(coun))] = max(start_row_df + len(df), len(dfper))
        d["max_col {0}".format(str(coun))] = len(df.columns) + 1

wb = openpyxl.load_workbook(output_directory)
for i in range(1, coun+1):
    sheet = wb[d["sheet " + str(i)]]
    max_row = d["max_row " + str(i)]
    max_col = d["max_col " + str(i)]
    for row in range(start_row_df + 1, max_row + 2):
        for col in range(1, max_col + 1):
            sheet.cell(row=start_row_df + 1, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True)
            sheet.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='left')

        col_widths = [19, 11, 12, 10.5, 16.5, 14, 11, 11, 19.2, 13.3, 7.5, 10, 10,10,10]
        for i, width in enumerate(col_widths, start=1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

wb.save(output_directory)
end_time = time.time()
# Calculate and print the elapsed time
elapsed_time = end_time - start_time
print(f"Total processing time for {coun} patient folder: {elapsed_time:.2f} seconds")